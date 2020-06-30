from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import time
import random
from selenium import webdriver
from selenium.webdriver.support import ui
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import sys
import datetime


def page_is_loaded(driver):
    return driver.find_element_by_tag_name("body") != None

def modal_page_is_loaded(driver):
    return driver.find_element_by_id("modal-body") != None

def normalize_excel_sheet():
    # Give the location of the file 
    loc = ("./contacts.xlsx") 

    # To open Workbook 
    wb = load_workbook(filename = loc)
    ws = wb.active 

    # initialize variables
    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

    lightredFill = PatternFill(start_color='ffcccb',
                   end_color='ffcccb',
                   fill_type='solid')

    
    # delete rows that don't have a phone number
    for row in ws:
        if not any(cell.value for cell in row):
            # print('empty row', row[0].row)
            ws.delete_rows(row[0].row, 1)

    # split full name into first and last
    row_count = ws.max_row
    print('row count', row_count)

    for row in ws.iter_rows(min_row = 2, min_col = 1, max_col = 1, max_row = row_count):
        for cell in row:
            name = cell.value
            try:
                name.strip()
            except:
                continue

            spaces = 0
            for char in name:
                if char == ' ':
                    spaces += 1

            # if no phone number
            if cell.offset(0,1).value == None:
                cell.fill = lightredFill
                cell.offset(0,1).fill = redFill
                cell.offset(0,2).fill = lightredFill
                cell.offset(0,3).fill = lightredFill
                cell.offset(0,4).fill = lightredFill
                cell.offset(0,5).fill = lightredFill
                cell.offset(0,6).fill = lightredFill

            if spaces == 0:
                cell.offset(0,3).value = name

            elif spaces == 1:
                first, last = name.split()
                cell.offset(0,3).value = first
                cell.offset(0,4).value = last

            elif spaces == 2:
                first, middle, last = name.split()
                cell.offset(0,3).value = first + " " + middle
                cell.offset(0,4).value = last

            else:
                cell.fill = redFill
                cell.offset(0,1).fill = lightredFill
                cell.offset(0,2).fill = lightredFill
                cell.offset(0,3).fill = lightredFill
                cell.offset(0,4).fill = lightredFill
                cell.offset(0,5).fill = lightredFill
                cell.offset(0,6).fill = lightredFill

    # format phone numbers to exclude special characters
    for row in ws.iter_rows(min_row = 2, min_col = 2, max_col = 2, max_row = row_count):
        for cell in row:
            number = cell.value
            try:
                number.strip()
            except:
                continue
            
            # remove anything that isn't a number from the string
            formatted_num = ''
            for char in number:
                if char.isdigit():
                    formatted_num += str(char)

            # if not 10 numbers, flag that cell
            if len(formatted_num) != 10:
                cell.fill = redFill
                cell.offset(0,-1).fill = lightredFill
                cell.offset(0,1).fill = lightredFill
                cell.offset(0,2).fill = lightredFill
                cell.offset(0,3).fill = lightredFill
                cell.offset(0,4).fill = lightredFill
                cell.offset(0,5).fill = lightredFill

            else:
                cell.offset(0,4).value = formatted_num

    for row in ws.iter_rows(min_row = 2, min_col = 3, max_col = 3, max_row = row_count):
        for cell in row:
            gender = str(cell.value)

        if gender.lower() == 'male' or gender.lower() == 'm' or gender.lower() == 'boy' or gender.lower() == 'guy':
            cell.offset(0,4).value = 'male'

        elif gender.lower() == 'female' or gender.lower() == 'f' or gender.lower() == 'girl' or gender.lower() == 'gal':
            cell.offset(0,4).value = 'female'

        elif cell.offset(0, -1).value != None and gender.lower() == 'none':
            cell.fill = redFill
            cell.offset(0,-1).fill = lightredFill
            cell.offset(0,-2).fill = lightredFill
            cell.offset(0,1).fill = lightredFill
            cell.offset(0,2).fill = lightredFill
            cell.offset(0,3).fill = lightredFill
            cell.offset(0,4).fill = lightredFill

        elif cell.offset(0, -1).value == None and gender.lower() == 'none':
            continue

        else:
            cell.offset(0,4).value = 'other'

    wb.save(filename = './contacts_formatted.xlsx')

def get_contact_list():
    loc = ('./contacts_formatted.xlsx') 

    # open Workbook 
    wb = load_workbook(filename = loc)
    ws = wb.active 

    row_count = ws.max_row

    all_contacts = []
    for row in ws.iter_rows(min_row = 2, min_col = 4, max_col = 7, max_row = row_count):
        single_contact_info = []
        for cell in row:

            # if the cell is red then there is a problem with that person's info, don't add it to the list
            if cell.fill.start_color.index == '00ffcccb':
                break
            single_contact_info.append(cell.value)
            if cell.column == 7 and single_contact_info != [None, None, None, None]:
                all_contacts.append(single_contact_info)
            
    print(all_contacts)
    return all_contacts

def find_labels():
    ask = False

    with open('./labels.txt') as f:
        labels = f.readlines()

    #remove whitespace characters like `\n` at the end of each line
    labels = [x.strip().lower() for x in labels]

    for x in labels:
        print('x', x, 'labels', labels)
        for char in range(0,4):
            if x[char].isdigit():
                continue
            else:
                if not ask:
                    ask = ask_to_follow_convention(x)

        if x[5:9].lower() == 'spri' or x[5:9].lower() == 'fall' or x[5:9].lower() == 'wint' or x[5:9].lower() == 'summ':
            continue
        else:
            if not ask:
                ask = ask_to_follow_convention(x)

    return labels

def ask_to_follow_convention(x):
    cont = 'change me'
    while cont.lower() != 'y' and cont.lower() != 'n':
        cont = input('\n\n\n\nThe item "' + x + '" in the "labels.txt" file does NOT meet the established naming convention for labels (found in labels_convention). Would you like to continue anyways? (y/n)       ')
    if cont.lower() == 'n':
        print('\n\nThanks for keeping missionhub organized ;)')
        time.sleep(1)
        sys.exit()
    else:
        sure = 'change me'
        while sure.lower() != 'y' and sure.lower() != 'n':
            print('sure', sure)
            sure = input('\n\nThis may result in missionhub becoming unorganized. Are you sure you want to continue? (y/n)        ')
        if sure.lower() == 'n':
            print('\n\nThanks for keeping missionhub organized ;)')
            time.sleep(1)
            sys.exit()

    return True

def close_blank_page(driver, wait, link):
    # open webpage
    driver.get(link)
    wait.until(page_is_loaded)

    # close the blank page that opens default with selenium and assign a main window 
    windows = driver.window_handles
    for window in windows:
        driver.switch_to.window(window)
        if len(driver.find_elements_by_css_selector("*")) >= 10:
            main_window = window
        else:
            driver.switch_to.window(window)
            driver.close()
    driver.switch_to.window(main_window)

    return main_window

def login_to_missionhub(driver, wait, main):
    time.sleep(1)

    driver.find_element_by_xpath('//*[@id="menu-item-1971"]/a').click()
    wait.until(page_is_loaded)

    driver.find_element_by_xpath('/html/body/ui-view/app/section/ui-view/sign-in/div/p[2]/a').click()
    wait.until(page_is_loaded)

    windows = driver.window_handles
    driver.switch_to.window(windows[-1])

    driver.find_element_by_xpath('//*[@id="email"]').send_keys('connorivy15@gmail.com')
    driver.find_element_by_xpath('//*[@id="pass"]').send_keys('September15!')
    driver.find_element_by_xpath('//*[@id="u_0_0"]').click()

    driver.switch_to.window(main)
    wait.until(page_is_loaded)
    time.sleep(5)

def add_new_contact(driver, wait, contact_info, user_labels):
    driver.find_element_by_xpath('/html/body/ui-view/app/section/ui-view/my-people-dashboard/div/div[1]/organization/accordion/div[1]/accordion-header/div/div[2]/ng-md-icon[1]').click()
    wait.until(page_is_loaded)

    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[6]/div[1]/label/input').send_keys(contact_info[0])
    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[6]/div[2]/label/input').send_keys(contact_info[1])
    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[6]/div[5]/div/label/div[2]/input').send_keys(contact_info[2])

    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[3]/label/assigned-people-select/div/div[1]/span/span/span/span[1]').click()
    
    # male 
    # /html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[6]/div[3]/label[1]/input

    # add label button
    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[1]/person-profile/form/div[1]/div[1]/div[1]/ng-md-icon').click()
    time.sleep(.75)
    availible_labels = driver.find_element_by_xpath('//*[@id="modal-body"]/multiselect-list/ul')

    a = datetime.datetime.now()
    list_elements = availible_labels.find_elements_by_xpath('.//*')
    for child in range (0,len(list_elements),3):  
        if list_elements[child].text.lower() in user_labels:
            list_elements[child].find_element_by_css_selector('span[class=ng-binding]').click()

    b = datetime.datetime.now()

    print('time', b-a)

    # the OK btn
    driver.find_element_by_xpath('/html/body/div[1]/div/div/edit-group-or-label-assignments/div[3]/button[2]/span').click()
    time.sleep(.75)

    # save btn
    driver.find_element_by_xpath('/html/body/div[1]/div/div/person-page/async-content/div/div[2]/button').click()
    wait.until(page_is_loaded)


def main():
    
    chromedriver = "chromedriver.exe"
    driver = webdriver.Chrome(chromedriver)
    wait = ui.WebDriverWait(driver, 10)
    link = 'https://get.missionhub.com/'

    # normalize_excel_sheet()
    contact_list = get_contact_list()
    labels = find_labels()
    main_window = close_blank_page(driver, wait, link)
    login_to_missionhub(driver, wait, main_window)

    for contact in contact_list:
        add_new_contact(driver, wait, contact, labels)

    time.sleep(60)


if __name__ == "__main__":
    main()