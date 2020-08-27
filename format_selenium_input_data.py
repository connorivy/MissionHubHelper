from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime
import time
import sys

col_num_min_formatted = 0
col_num_max = 0

def normalize_excel_sheet():
    global col_num_min_formatted
    global col_num_max
    # Give the location of the file 
    loc = ("./supporting_files/contacts.xlsx") 

    # To open Workbook 
    wb = load_workbook(filename = loc)
    ws = wb.active 

    # initialize variables
    white = PatternFill(start_color='FFFFFF',
                   end_color='FFFFFF',
                   fill_type='solid')

    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

    lightredFill = PatternFill(start_color='ffcccb',
                   end_color='ffcccb',
                   fill_type='solid')

    # define the location of each column
    col_num_full_name = 1
    col_num_phone_num = 2
    col_num_gender = 3
    col_num_year = 4
    col_num_notes = 5

    col_num_min_formatted = 5

    col_num_first_name = 6
    col_num_last_name = 7
    col_num_phone_num_formatted = 8
    col_num_gender_formatted = 9
    col_num_year_formatted = 10

    col_num_max = 10

    row_count = ws.max_row
    # format names split into first and last
    for row in ws.iter_rows(min_row = 2, min_col = col_num_first_name, max_col = col_num_last_name, max_row = row_count):
        for cell in row:
            cell.fill = white
            f_l_name = cell.value

            if isinstance(f_l_name, str):
                cell.value = f_l_name.strip()

    # split full name into first and last
    for row in ws.iter_rows(min_row = 2, min_col = col_num_full_name, max_col = col_num_full_name, max_row = row_count):
        for cell in row:
            cell.fill = white
            
            name = cell.value
            if name == None:
                continue

            if isinstance(name, str):
                name = name.strip()
            else:
                name = str(name)

            spaces = 0
            for char in name:
                if char == ' ':
                    spaces += 1

            if spaces == 0:
                cell.offset(0,col_num_first_name - col_num_full_name).value = name

            elif spaces == 1:
                first, last = name.split()
                cell.offset(0,col_num_first_name - col_num_full_name).value = first
                cell.offset(0,col_num_last_name - col_num_full_name).value = last

            elif spaces == 2:
                first, middle, last = name.split()
                cell.offset(0,col_num_first_name - col_num_full_name).value = first + " " + middle
                cell.offset(0,col_num_last_name - col_num_full_name).value = last

            else:
                flag_cell(cell,col_num_full_name,redFill,lightredFill)

    # format phone numbers to exclude special characters
    for row in ws.iter_rows(min_row = 2, min_col = col_num_phone_num, max_col = col_num_phone_num, max_row = row_count):
        for cell in row:
            cell.fill = white
            number = cell.value

            # if no phone number, flag the cell and move to the next number
            if number == None:
                flag_cell(cell,col_num_phone_num,redFill, lightredFill)
                continue

            # remove anything that isn't a number from the string
            if isinstance(number, str):
                number.strip()
            else:
                number = str(number)

            formatted_num = ''
            for char in number:
                if char.isdigit():
                    formatted_num += str(char)

            # if not 10 numbers, flag that cell
            if len(formatted_num) != 10:
                flag_cell(cell,col_num_phone_num,redFill, lightredFill)
                continue

            else:
                cell.offset(0,col_num_phone_num_formatted - col_num_phone_num).value = formatted_num

    # format gender
    for row in ws.iter_rows(min_row = 2, min_col = col_num_gender, max_col = col_num_gender, max_row = row_count):
        for cell in row:
            cell.fill = white
            gender = cell.value
            
            if gender == None:
                continue
            if isinstance(gender, str):
                gender.strip()
            else:
                gender = str(gender)

            if gender.lower() == 'male' or gender.lower() == 'm' or gender.lower() == 'boy' or gender.lower() == 'guy':
                cell.offset(0,col_num_gender_formatted - col_num_gender).value = 'male'

            elif gender.lower() == 'female' or gender.lower() == 'f' or gender.lower() == 'girl' or gender.lower() == 'gal':
                cell.offset(0,col_num_gender_formatted - col_num_gender).value = 'female'

            else:
                cell.offset(0,col_num_gender_formatted - col_num_gender).value = 'other'

    # format year
    for row in ws.iter_rows(min_row = 2, min_col = col_num_year, max_col = col_num_year, max_row = row_count):
        for cell in row:
            cell.fill = white
            
            year = cell.value
            if year == None:
                continue

            if isinstance(year, str):
                year = year.strip()
            else:
                year = str(year)

            now = datetime.datetime.now()
            current_year = now.year
            month = now.month
            freshman_year = 0000
            if month > 7:
                freshman_year = current_year + 4
            elif month < 5:
                freshman_year = current_year + 3
            else:
                while freshman_year < current_year - 10 or freshman_year > current_year + 10:
                    try:
                        freshman_year = int(input('What year will these freshmen graduate?      '))
                    except:
                        continue

            if year.lower() == 'first' or year.lower() == 'f' or 'fr' in year.lower() or year.lower() == '1st' or year.lower() == '1':
                cell.offset(0,col_num_year_formatted - col_num_year).value = str(freshman_year)
            elif year.lower() == 'second' or 'so' in year.lower() or year.lower() == '2nd' or year.lower() == '2':
                cell.offset(0,col_num_year_formatted - col_num_year).value = str(freshman_year-1)
            elif year.lower() == 'third' or year.lower() == 'j' or 'ju' in year.lower() or year.lower() == '3rd' or year.lower() == '3':
                cell.offset(0,col_num_year_formatted - col_num_year).value = str(freshman_year-2)
            elif year.lower() == 'fourth' or year.lower() == 'senior' or year.lower() == '4th' or year.lower() == '5':
                cell.offset(0,col_num_year_formatted - col_num_year).value = str(freshman_year-3)
            

    wb.save(filename = './supporting_files/contacts_formatted_do_not_edit.xlsx')

def flag_cell(cell, col_num, redFill, lightredFill):
    for x in range(1, 10):
        cell_offset = x - col_num
        if cell_offset == 0:
            cell.offset(0,cell_offset).fill = redFill
        else:
            cell.offset(0,cell_offset).fill = lightredFill

def get_contact_list():
    loc = ('./supporting_files/contacts_formatted_do_not_edit.xlsx') 

    # open Workbook 
    wb = load_workbook(filename = loc)
    ws = wb.active 

    row_count = ws.max_row

    all_contacts = []
    counter = 1
    for row in ws.iter_rows(min_row = 2, min_col = col_num_min_formatted, max_col = col_num_max, max_row = row_count):
        single_contact_info = []
        counter += 1
        for cell in row:

            # if the cell is red then there is a problem with that person's info, don't add it to the list
            if cell.fill.start_color.index == '00ffcccb':
                print('***** Problem adding row number', counter, '*****')
                break
            single_contact_info.append(cell.value)
            if cell.column == col_num_max and single_contact_info != [None, None, None, None, None, None]:
                all_contacts.append(single_contact_info)
            
    return all_contacts

def find_labels():
    ask = False

    with open('./supporting_files/labels.txt') as f:
        labels = f.readlines()

    # remove whitespace characters like `\n` at the end of each line
    labels = [x.strip().lower() for x in labels]

    for x in labels:
        for char in range(0,4):
            if x[char].isdigit():
                continue
            else:
                if not ask:
                    ask = ask_to_follow_convention(x)

        if x[5:9].lower() == 'spri' or x[5:9].lower() == 'fall' or x[5:9].lower() == 'wint' or x[5:9].lower() == 'summ':
            continue
        else:
            if not ask:
                ask = ask_to_follow_convention(x)

    return labels

def ask_to_follow_convention(x):
    cont = 'change me'
    print('\n\n\n')
    while cont.lower() != 'y' and cont.lower() != 'n':
        cont = input('\nThe item "' + x + '" in the "labels.txt" file does NOT meet the established naming convention for labels (found in supporting_files/labels_convention). Would you like to continue anyways? (y/n)       ')
    if cont.lower() == 'n':
        print('\n\nThanks for keeping missionhub organized ;)')
        time.sleep(1)
        sys.exit()
    else:
        sure = 'change me'
        while sure.lower() != 'y' and sure.lower() != 'n':
            sure = input('\n\nThis may result in missionhub becoming unorganized. Are you sure you want to continue? (y/n)        ')
        if sure.lower() == 'n':
            print('\n\nThanks for keeping missionhub organized ;)')
            time.sleep(1)
            sys.exit()

    return True
